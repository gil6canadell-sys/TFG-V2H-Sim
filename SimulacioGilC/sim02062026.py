"""
SIMULACIÓ V2H — MULTI-ESCENARI (VERSIÓ CORREGIDA)
==================================================================
Autor: Gil Canadell Montaner — TFG ETSEIB/UPC
Data: 2026
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import os
import datetime

FITXER_ENTRADA = "Model_V2H_v7_8760h_fix.xlsx"
FITXER_CONSUMS = "ConsumsAnuals.xlsx"
FITXER_FV_LAB = "FV_Laborable_Ordenat.xlsx"
FITXER_FV_FDS = "FV_Findes_Ordenat.xlsx"
FITXER_SORTIDA = "Resultats_Totals_V2H.xlsx"
FULLA_BASE = "Sim8760"

escenaris_habitatge = [
    {"nom": "PIS",       "FV_activada": False, "tipus_consum": "PIS"},
    {"nom": "CASA_FV",   "FV_activada": True,  "tipus_consum": "CASA"},
    {"nom": "CASA_noFV", "FV_activada": False, "tipus_consum": "CASA"}
]

escenaris_bateria = [
    {"nom": "Bat60", "C_bat": 60, "SOC_max": 0.90, "SOC_min": 0.30,
     "P_carg": 7.4, "P_desc": 7.4, "eta_c": 0.95, "eta_d": 0.95}
]

escenaris_mobilitat = [
    {"nom": "Esc1_Alt", "d_lab": 40, "d_fds": 30, "t_sort_lab": 7,  "t_arr_lab": 19, "t_sort_fds": 10, "t_arr_fds": 15, "c_esp": 0.18},
    {"nom": "Esc2_Mod", "d_lab": 30, "d_fds": 20, "t_sort_lab": 7,  "t_arr_lab": 19, "t_sort_fds": 10, "t_arr_fds": 14, "c_esp": 0.18},
    {"nom": "Esc3_Bai", "d_lab": 15, "d_fds": 15, "t_sort_lab": 7,  "t_arr_lab": 19, "t_sort_fds": 10, "t_arr_fds": 13, "c_esp": 0.18}
]

escenaris_carrega_ext = [
    {"nom": "Base",  "carrega_ext": "sense",        "tarifa_ext": 0.0},
    {"nom": "CFein", "carrega_ext": "feina_gratis",  "tarifa_ext": 0.0},
    {"nom": "CPubl", "carrega_ext": "publica",       "tarifa_ext": 0.40},
]

MARGE_SOC = 0.05
COST_CARREGADOR_BIDIR = 5000
COST_CARREGADOR_UNIDIR = 1350
TARIFA_CARRER = 0.40
COST_FIX_MENSUAL = 19

CICLES_VIDA_BATERIA = 3000
PREU_BATERIA_USD_KWH = 81      # BloombergNEF Battery Price Survey 2025
USD_TO_EUR = 0.92
PREU_BATERIA_EUR_KWH = PREU_BATERIA_USD_KWH * USD_TO_EUR
DEGRADACIO_SOH_TOTAL = 0.20

PREU_EXCEDENT = {
    'P1': 0.051135,
    'P2': 0.052685,
    'P3': 0.029769,
}

def periode_tarifari(hora, fds):
    """Retorna el període tarifari (P1/P2/P3) segons tarifa 2.0TD.
    - P1 (punta): 10-14h i 18-22h (laborables)
    - P2 (pla):   8-10h, 14-18h i 22-24h (laborables)
    - P3 (vall):  0-8h (laborables) + tot el cap de setmana
    """
    if fds == 1:
        return 'P3'
    if 10 <= hora < 14 or 18 <= hora < 22:
        return 'P1'
    elif 8 <= hora < 10 or 14 <= hora < 18 or 22 <= hora < 24:
        return 'P2'
    else:
        return 'P3'

def simular(wb, conf):
    """Executa la simulació horària anual (8760h) d'un escenari V2H.
    
    Paràmetres (via diccionari 'conf'):
        - Bateria: C_bat, SOC_max, SOC_min, P_carg, P_desc, eta_c, eta_d
        - Mobilitat: d_lab, d_fds, c_esp, t_sort/arr_lab/fds
        - Habitatge: FV_activada, llista_consums, llista_produccio_fv
        - Economia: llista_tarifes, llista_tco2, carrega_ext, tarifa_ext
    
    Retorna:
        resultats: llista de 8760 diccionaris amb dades horàries
        resum: diccionari amb totals anuals (energètics, econòmics, ambientals)
    """
    C_bat = conf["C_bat"]
    SOC_max = conf["SOC_max"]
    SOC_min = conf["SOC_min"]
    P_carg = conf["P_carg"]
    P_desc = conf["P_desc"]
    eta_c = conf["eta_c"]
    eta_d = conf["eta_d"]
    
    d_lab = conf["d_lab"]
    d_fds = conf["d_fds"]
    c_esp = conf["c_esp"]
    t_sort_lab = conf["t_sort_lab"]
    t_arr_lab = conf["t_arr_lab"]
    t_sort_fds = conf["t_sort_fds"]
    t_arr_fds = conf["t_arr_fds"]
    FV_activada = conf["FV_activada"]
    llista_consums = conf["llista_consums"]
    llista_tarifes = conf["llista_tarifes"]
    llista_tco2 = conf["llista_tco2"]
    llista_produccio_fv = conf["llista_produccio_fv"]

    carrega_ext = conf.get("carrega_ext", "sense")
    tarifa_ext = conf.get("tarifa_ext", 0.0)

    # Amb càrrega externa només cal reservar SOC per l'anada (meitat del trajecte)
    if carrega_ext in ("feina_gratis", "publica"):
        soc_res_lab = SOC_min + (d_lab * c_esp / 2) / C_bat + MARGE_SOC
    else:
        soc_res_lab = SOC_min + (d_lab * c_esp) / C_bat + MARGE_SOC
    
    soc_res_fds = SOC_min + (d_fds * c_esp) / C_bat + MARGE_SOC

    t_full_lab = t_sort_lab - 5
    t_full_fds = t_sort_fds - 5
    h_rampa_lab = (24 - t_arr_lab) + t_full_lab
    h_rampa_fds = (24 - t_arr_fds) + t_full_fds

    def calc_soc_limit(hora, fds, delta):
        if delta == 0: return SOC_min
        if fds == 0:  
            if hora >= t_arr_lab:
                frac = (hora - t_arr_lab) / h_rampa_lab
                return SOC_min + (soc_res_lab - SOC_min) * frac
            elif hora < t_full_lab:
                frac = ((24 - t_arr_lab) + hora) / h_rampa_lab
                return SOC_min + (soc_res_lab - SOC_min) * frac
            elif hora < t_sort_lab: return soc_res_lab
            else: return SOC_min
        else:  
            if hora >= t_arr_fds:
                frac = (hora - t_arr_fds) / h_rampa_fds
                return SOC_min + (soc_res_fds - SOC_min) * frac
            elif hora < t_full_fds:
                frac = ((24 - t_arr_fds) + hora) / h_rampa_fds
                return SOC_min + (soc_res_fds - SOC_min) * frac
            elif hora < t_sort_fds: return soc_res_fds
            else: return SOC_min

    ws = wb[conf["fulla_dades"]]
    soc = SOC_max
    prev_delta = 1
    
    resultats = []
    total_v2h = total_carg = total_xarxa = total_estalvi = total_cost_xarxa = 0
    total_carg_ext = 0
    total_cost_carg_ext = 0
    total_estalvi_unidir = 0
    total_estalvi_fv = 0
    total_excedent = 0
    total_ingres_excedent = 0
    total_carg_fv_ve = 0
    total_estalvi_fv_ve = 0
    soc_min_year = 1.0
    below_soc_min = 0
    
    for idx, r in enumerate(range(4, 8764)):
        hora_raw = ws.cell(r, 3).value
        hora = (hora_raw - 1) % 24  # 1-24 → 0-23 (hora=25 canvi horari → 0)
        fds = ws.cell(r, 4).value
        data = ws.cell(r, 2).value

        tarifa_raw = llista_tarifes[idx]
        tarifa = tarifa_raw if tarifa_raw > 0 else 0.09

        if fds == 0:
            delta = 0 if t_sort_lab <= hora < t_arr_lab else 1
        else:
            delta = 0 if t_sort_fds <= hora < t_arr_fds else 1

        ws.cell(r, 5).value = delta

        p_llar = llista_consums[idx]
        ws.cell(r, 6).value = round(p_llar, 3)

        if FV_activada:
            p_fv = llista_produccio_fv[idx]
        else:
            p_fv = 0
            
        p_net = p_llar - p_fv
        soc_limit = calc_soc_limit(hora, fds, delta)

        soc_inici = soc
        dist = d_fds if fds else d_lab
        
        if delta == 0 and prev_delta == 1:
            soc_inici -= (dist * c_esp / 2) / C_bat
        elif delta == 1 and prev_delta == 0:
            soc_inici -= (dist * c_esp / 2) / C_bat

        soc_inici = max(0, soc_inici)

        e_carg_ext = 0
        if delta == 0 and fds == 0 and carrega_ext != "sense":
            espai_disponible = max(0, (SOC_max - soc_inici) * C_bat)
            e_carg_ext = min(P_carg, espai_disponible) * eta_c
        
        if delta == 0:
            action, e_carg, e_v2h = "Fora", 0, 0
            if e_carg_ext > 0:
                action = "Carg ext"
        elif soc_inici < soc_limit:
            action, e_carg, e_v2h = "Carg xar", min(P_carg, max(0, (SOC_max - soc_inici) * C_bat)) * eta_c, 0
        elif p_net < 0 and soc_inici < SOC_max:
            action, e_carg, e_v2h = "Carg FV", min(abs(p_net), P_carg, max(0, (SOC_max - soc_inici) * C_bat)) * eta_c, 0
        elif p_net > 0 and soc_inici > soc_limit:
            action, e_carg, e_v2h = "V2H", 0, min(max(0, p_net), P_desc, max(0, (soc_inici - soc_limit) * C_bat)) * eta_d
        elif soc_inici < SOC_max:
            action, e_carg, e_v2h = "Carg xar", min(P_carg, max(0, (SOC_max - soc_inici) * C_bat)) * eta_c, 0
        else:
            action, e_carg, e_v2h = "Stby", 0, 0
        
        carg_fv_ve_hora = 0
        if delta == 0:
            e_xarxa = max(0, p_net)
        else:
            if action == "Carg xar":
                if p_net < 0:
                    # Excedent FV: primer aprofitem el surplus per carregar el VE
                    fv_disponible = abs(p_net)
                    necessitat_carg = e_carg / eta_c
                    fv_per_carregar = min(fv_disponible, necessitat_carg)
                    carg_fv_ve_hora = fv_per_carregar * eta_c
                    e_xarxa = max(0, necessitat_carg - fv_per_carregar)
                else:
                    e_xarxa = max(0, p_net) + (e_carg / eta_c)
            elif action == "Carg FV":
                e_xarxa = 0
                carg_fv_ve_hora = e_carg
            elif action == "V2H":
                e_xarxa = max(0, p_net - e_v2h)
            else:
                e_xarxa = max(0, p_net)

        soc_fi = soc_inici + (e_carg / C_bat) + (e_carg_ext / C_bat)
        if action == "V2H":
            soc_fi -= e_v2h / (eta_d * C_bat)

        soc_fi = min(SOC_max, soc_fi)

        if soc_fi < SOC_min: below_soc_min += 1
        if soc_fi < soc_min_year: soc_min_year = soc_fi

        estalvi_v2h = e_v2h * tarifa
        cost_xarxa_hora = e_xarxa * tarifa
        cost_ext_hora = (e_carg_ext / eta_c) * tarifa_ext

        # Estalvi unidir: diferència entre carregar al carrer vs a casa
        if e_carg > 0:
            cost_carrer_hora = (e_carg / eta_c) * TARIFA_CARRER
            if action == "Carg FV":
                cost_casa_hora = 0
            elif action == "Carg xar" and carg_fv_ve_hora > 0:
                kwh_de_xarxa = (e_carg - carg_fv_ve_hora) / eta_c
                cost_casa_hora = max(0, kwh_de_xarxa) * tarifa
            else:
                cost_casa_hora = (e_carg / eta_c) * tarifa
            estalvi_unidir_hora = cost_carrer_hora - cost_casa_hora
        else:
            estalvi_unidir_hora = 0
        
        autoconsum_fv_hora = min(p_fv, p_llar) if FV_activada else 0
        estalvi_fv_hora = autoconsum_fv_hora * tarifa

        if carg_fv_ve_hora > 0:
            estalvi_fv_ve_hora = (carg_fv_ve_hora / eta_c) * tarifa
        else:
            estalvi_fv_ve_hora = 0

        if FV_activada and p_net < 0:
            sobrant_fv = abs(p_net)
            if action == "Carg FV":
                excedent_hora = max(0, sobrant_fv - (e_carg / eta_c))
            elif action == "Carg xar" and carg_fv_ve_hora > 0:
                excedent_hora = max(0, sobrant_fv - (carg_fv_ve_hora / eta_c))
            else:
                excedent_hora = sobrant_fv
        else:
            excedent_hora = 0
        
        periode = periode_tarifari(hora, fds)
        ingres_excedent_hora = excedent_hora * PREU_EXCEDENT[periode]
        
        resultats.append({
            'fila': r, 'data': data, 'hora': hora, 'fds': fds, 'delta': delta,
            'p_llar': p_llar, 'p_fv': p_fv, 'p_net': p_net, 'soc_limit': soc_limit,
            'soc_inici': soc_inici, 'action': action, 'e_carg': e_carg, 'e_v2h': e_v2h,
            'e_xarxa': e_xarxa, 'soc_fi': soc_fi, 'tarifa': tarifa, 'estalvi': estalvi_v2h,
            'e_carg_ext': e_carg_ext, 'cost_ext': cost_ext_hora,
        })
        
        total_v2h += e_v2h
        total_carg += e_carg
        total_xarxa += e_xarxa
        total_estalvi += estalvi_v2h
        total_cost_xarxa += cost_xarxa_hora
        total_carg_ext += e_carg_ext
        total_cost_carg_ext += cost_ext_hora
        total_estalvi_unidir += estalvi_unidir_hora
        total_estalvi_fv += estalvi_fv_hora
        total_excedent += excedent_hora
        total_ingres_excedent += ingres_excedent_hora
        total_carg_fv_ve += carg_fv_ve_hora
        total_estalvi_fv_ve += estalvi_fv_ve_hora

        prev_delta = delta
        soc = soc_fi
        
    xarxa_diari = {}
    for idx_r, res in enumerate(resultats):
        dia = idx_r // 24
        xarxa_diari[dia] = xarxa_diari.get(dia, 0) + res['e_xarxa']
    
    co2_diari = {}
    for dia in sorted(xarxa_diari.keys()):
        tco2_kwh = llista_tco2[dia] if dia < len(llista_tco2) else 0.0
        co2_diari[dia] = xarxa_diari[dia] * tco2_kwh
    
    total_co2 = sum(co2_diari.values())

    estalvi_net = total_estalvi + total_estalvi_fv_ve - total_cost_carg_ext
    cost_fix_anual = COST_FIX_MENSUAL * 12

    payback_unidir = COST_CARREGADOR_UNIDIR / total_estalvi_unidir if total_estalvi_unidir > 0 else 999

    diferencia_cost = COST_CARREGADOR_BIDIR - COST_CARREGADOR_UNIDIR
    payback_diferencial = diferencia_cost / estalvi_net if estalvi_net > 0 else 999

    # Degradació: (EFC / cicles_vida) × %SOH_perdut × cost_bateria
    efc_v2h = total_v2h / (2 * C_bat)
    pct_degradacio = efc_v2h * DEGRADACIO_SOH_TOTAL / CICLES_VIDA_BATERIA
    cost_bateria_total = C_bat * PREU_BATERIA_EUR_KWH
    cost_degradacio_anual = pct_degradacio * cost_bateria_total

    estalvi_net_amb_deg = estalvi_net - cost_degradacio_anual
    payback_bidir_amb_deg = COST_CARREGADOR_BIDIR / estalvi_net_amb_deg if estalvi_net_amb_deg > 0 else 999
    payback_difer_amb_deg = diferencia_cost / estalvi_net_amb_deg if estalvi_net_amb_deg > 0 else 999

    resum = {
        'total_v2h': total_v2h, 'total_carg': total_carg, 'total_xarxa': total_xarxa,
        'total_estalvi': total_estalvi, 'estalvi_net': estalvi_net,
        'total_estalvi_fv': total_estalvi_fv,
        'total_excedent': total_excedent,
        'total_ingres_excedent': total_ingres_excedent,
        'estalvi_total': total_estalvi + total_estalvi_fv + total_estalvi_fv_ve + total_ingres_excedent,
        'soc_min_year': soc_min_year,
        'below_soc_min': below_soc_min, 'efc_v2h': efc_v2h,
        'payback': COST_CARREGADOR_BIDIR / estalvi_net if estalvi_net > 0 else 999,
        'payback_unidir': payback_unidir,
        'payback_diferencial': payback_diferencial,
        'cost_degradacio_anual': cost_degradacio_anual,
        'payback_bidir_amb_deg': payback_bidir_amb_deg,
        'payback_difer_amb_deg': payback_difer_amb_deg,
        'total_carg_fv_ve': total_carg_fv_ve,
        'total_estalvi_fv_ve': total_estalvi_fv_ve,
        'estalvi_anual_unidir': total_estalvi_unidir,
        'soc_final': soc,
        'total_llar': sum(r['p_llar'] for r in resultats),
        'factura_anual': total_cost_xarxa + cost_fix_anual - total_ingres_excedent,
        'total_co2': total_co2,
        'co2_diari': co2_diari,
        'total_carg_ext': total_carg_ext,
        'total_cost_carg_ext': total_cost_carg_ext,
        'cost_total_energia': total_cost_xarxa + total_cost_carg_ext + cost_fix_anual - total_ingres_excedent,
    }
    
    return resultats, resum

def escriure_resultats(wb, resultats, resum, fulla):
    """Escriu els resultats horaris i totals anuals a la fulla Excel de l'escenari."""
    ws = wb[fulla]
    font_dades = Font(name='Calibri', size=9)
    vora = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    centrat = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for res in resultats:
        r = res['fila']
        ws.cell(r, 7).value = round(res['p_fv'], 3)
        ws.cell(r, 8).value = round(res['p_net'], 3)
        ws.cell(r, 9).value = res['action']
        e_carg_mostrar = res['e_carg_ext'] if res['e_carg_ext'] > 0 else res['e_carg']
        ws.cell(r, 10).value = round(e_carg_mostrar, 3)
        ws.cell(r, 11).value = round(res['e_v2h'], 3)
        ws.cell(r, 12).value = round(res['e_xarxa'], 3)
        ws.cell(r, 13).value = round(res['soc_inici'], 4)
        ws.cell(r, 15).value = round(res['estalvi'], 4)
        ws.cell(r, 16).value = round(res['soc_limit'], 4)
        
        for c in [7, 8, 9, 10, 11, 12, 13, 15, 16]:
            cell = ws.cell(r, c)
            cell.font = font_dades
            cell.border = vora
            cell.alignment = centrat
    
    ws.cell(3, 17).value = "tCO2 diari"
    ws.cell(3, 17).font = Font(name='Calibri', size=9, bold=True)
    ws.cell(3, 17).border = vora
    ws.cell(3, 17).alignment = centrat
    
    co2_diari = resum['co2_diari']
    for dia_idx in sorted(co2_diari.keys()):
        fila_dia = 4 + dia_idx * 24
        cell = ws.cell(fila_dia, 17)
        cell.value = round(co2_diari[dia_idx], 8)
        cell.font = font_dades
        cell.border = vora
        cell.alignment = centrat
        cell.number_format = '0.000000'
    
    tr = 8764
    ws.cell(tr, 10).value = round(resum['total_carg'] + resum['total_carg_ext'], 2)
    ws.cell(tr, 11).value = round(resum['total_v2h'], 2)
    ws.cell(tr, 12).value = round(resum['total_xarxa'], 2)
    ws.cell(tr, 15).value = round(resum['total_estalvi'], 2)
    
    ws.cell(tr, 17).value = round(resum['total_co2'], 6)
    ws.cell(tr, 17).font = Font(name='Calibri', size=9, bold=True)
    ws.cell(tr, 17).border = vora
    ws.cell(tr, 17).alignment = centrat
    ws.cell(tr, 17).number_format = '0.000000'

if __name__ == "__main__":
    print("="*50)
    print("INICIANT SIMULACIÓ MULTI-ESCENARI V2H")
    print("="*50)
    
    if not os.path.exists(FITXER_CONSUMS):
        print(f"ERROR: No s'ha trobat '{FITXER_CONSUMS}'")
        exit()
        
    print(f"Extreient dades reals de consum de {FITXER_CONSUMS}...")
    wb_consums = load_workbook(FITXER_CONSUMS, data_only=True)
    ws_pis = wb_consums["ConsumPis25"]
    ws_casa = wb_consums["ConsumCasa25"]
    
    llista_pis = [float(ws_pis.cell(r, 4).value or 0) for r in range(2, 8762)]
    llista_casa = [float(ws_casa.cell(r, 4).value or 0) for r in range(2, 8762)]
    
    llista_tco2 = []
    for r in range(2, 367):  # 365 dies
        val = ws_pis.cell(r, 9).value
        llista_tco2.append(float(val) if val is not None else 0.0)
    
    llista_tarifes = []
    for r in range(2, 8762):
        val = ws_pis.cell(r, 6).value
        if val is not None:
            num = float(val)
            if num > 2.0:  # €/MWh → €/kWh
                num = num / 1000.0
            llista_tarifes.append(num)
        else:
            llista_tarifes.append(0.0)
    
    wb_consums.close()
    
    for f in [FITXER_FV_LAB, FITXER_FV_FDS]:
        if not os.path.exists(f):
            print(f"ERROR: No s'ha trobat '{f}'")
            exit()
    
    print(f"Carregant perfils FV de {FITXER_FV_LAB} i {FITXER_FV_FDS}...")
    
    def carregar_perfils_fv(fitxer):
        """Retorna dict amb claus 'auto' i 'total', cada una amb {mes: [24 valors]}"""
        wb_fv = load_workbook(fitxer, data_only=True)
        perfils = {'auto': {}, 'total': {}}
        for clau, nom_fulla in [('auto', 'Autoconsum FV (kWh)'), ('total', 'Producció Total FV (kWh)')]:
            ws_fv = wb_fv[nom_fulla]
            for mes in range(1, 13):
                col = mes + 1
                perfils[clau][mes] = []
                for fila in range(2, 26):  # 24 hores
                    val = ws_fv.cell(fila, col).value
                    perfils[clau][mes].append(float(val) if val else 0.0)
        wb_fv.close()
        return perfils
    
    fv_lab = carregar_perfils_fv(FITXER_FV_LAB)
    fv_fds = carregar_perfils_fv(FITXER_FV_FDS)
    
    llista_autoconsum_fv = []
    llista_produccio_fv = []
    
    dt = datetime.datetime(2025, 1, 1, 0, 0)
    for i in range(8760):
        mes = dt.month
        hora = dt.hour
        fds = 1 if dt.weekday() >= 5 else 0
        perfil = fv_fds if fds else fv_lab
        llista_autoconsum_fv.append(perfil['auto'][mes][hora])
        llista_produccio_fv.append(perfil['total'][mes][hora])
        dt += datetime.timedelta(hours=1)
    
    total_autoconsum = sum(llista_autoconsum_fv)
    total_produccio = sum(llista_produccio_fv)
    print(f"  -> Autoconsum FV anual: {total_autoconsum:.1f} kWh")
    print(f"  -> Producció Total FV anual: {total_produccio:.1f} kWh")
    
    consum_casa_original = sum(llista_casa)
    llista_casa = [llista_casa[i] + llista_autoconsum_fv[i] for i in range(8760)]
    consum_casa_corregit = sum(llista_casa)
    print(f"  -> Consum CASA original (compra xarxa): {consum_casa_original:.1f} kWh/any")
    print(f"  -> Consum CASA corregit (demanda real): {consum_casa_corregit:.1f} kWh/any")
    
    mitjana_tarifa = sum(llista_tarifes) / len(llista_tarifes) if llista_tarifes else 0
    print(f"  -> Dades carregades correctament.")
    print(f"  -> Mitjana tarifa detectada: {mitjana_tarifa:.4f} €/kWh")
    print(f"  -> tCO2/KWh diaris carregats: {len(llista_tco2)} dies (mitjana: {sum(llista_tco2)/len(llista_tco2):.6f} tCO2/KWh)")

    if not os.path.exists(FITXER_ENTRADA):
        print(f"ERROR: No s'ha trobat '{FITXER_ENTRADA}'")
        exit()

    print(f"Carregant l'arxiu base del model: {FITXER_ENTRADA}")
    wb = load_workbook(FITXER_ENTRADA)
    ws_base = wb[FULLA_BASE]

    nom_resum = "Resum_Global"
    if nom_resum in wb.sheetnames:
        del wb[nom_resum]
    ws_resum = wb.create_sheet(nom_resum, 0)
    
    ws_resum.append([
        "Escenari", "Estalvi V2H (€)", "Estalvi FV (€)", "Estalvi FV→VE (€)", "Ingrés Excedents (€)",
        "Estalvi Total V2H+FV+FV→VE+Exc (€)",
        "Cost Càrr. Ext. (€)", "Estalvi Net (€)",
        "Factura Xarxa (€)", "Cost Total Energia (€)",
        "V2H (kWh)", "Xarxa Total (kWh)", "Càrr. Ext. (kWh)", "Excedent FV (kWh)",
        "FV→VE (kWh)",
        "Demanda Llar (kWh)", "Autoconsum (%)", "EFC V2H",
        "Cost Degradació (€/any)",
        "Payback Bidir (anys)", "Payback Unidir (anys)", "Payback Diferencial (anys)",
        "PB Bidir c/Deg (anys)", "PB Difer c/Deg (anys)",
        "SOC min Any", "tCO2 Anual"
    ])
    for cell in ws_resum[1]: cell.font = Font(bold=True)

    total_simulacions = (len(escenaris_habitatge) * len(escenaris_bateria)
                         * len(escenaris_mobilitat) * len(escenaris_carrega_ext))
    comptador = 1
    
    resultats_csv = []
    
    for hab in escenaris_habitatge:
        for bat in escenaris_bateria:
            for mob in escenaris_mobilitat:
                for cext in escenaris_carrega_ext:
                    nom_escenari = f"{hab['nom']}_{bat['nom']}_{mob['nom']}_{cext['nom']}"
                    print(f"[{comptador}/{total_simulacions}] Simulant: {nom_escenari}")
                    
                    ws_nova = wb.copy_worksheet(ws_base)
                    ws_nova.title = nom_escenari
                    
                    consums_utilitzar = llista_pis if hab["tipus_consum"] == "PIS" else llista_casa
                    
                    conf_actual = {
                        "nom_hab": hab["nom"],
                        "fulla_dades": nom_escenari,
                        "FV_activada": hab["FV_activada"],
                        "llista_consums": consums_utilitzar,
                        "llista_tarifes": llista_tarifes,
                        "llista_tco2": llista_tco2,
                        "llista_produccio_fv": llista_produccio_fv,
                        "nom_bat": bat["nom"],
                        "nom_mob": mob["nom"],
                        "carrega_ext": cext["carrega_ext"],
                        "tarifa_ext": cext["tarifa_ext"],
                    }
                    conf_actual.update(bat)
                    conf_actual.update(mob)
                    
                    resultats, resum = simular(wb, conf_actual)
                    escriure_resultats(wb, resultats, resum, nom_escenari)
                    
                    autoconsum = (resum['total_v2h'] / (resum['total_v2h'] + resum['total_xarxa'])
                                  if (resum['total_v2h'] + resum['total_xarxa']) > 0 else 0)
                    
                    ws_resum.append([
                        nom_escenari,
                        round(resum['total_estalvi'], 2),
                        round(resum['total_estalvi_fv'], 2),
                        round(resum['total_estalvi_fv_ve'], 2),
                        round(resum['total_ingres_excedent'], 2),
                        round(resum['estalvi_total'], 2),
                        round(resum['total_cost_carg_ext'], 2),
                        round(resum['estalvi_net'], 2),
                        round(resum['factura_anual'], 2),
                        round(resum['cost_total_energia'], 2),
                        round(resum['total_v2h'], 1),
                        round(resum['total_xarxa'], 1),
                        round(resum['total_carg_ext'], 1),
                        round(resum['total_excedent'], 1),
                        round(resum['total_carg_fv_ve'], 1),
                        round(resum['total_llar'], 1),
                        round(autoconsum * 100, 2),
                        round(resum['efc_v2h'], 1),
                        round(resum['cost_degradacio_anual'], 2),
                        round(resum['payback'], 1),
                        round(resum['payback_unidir'], 1),
                        round(resum['payback_diferencial'], 1),
                        round(resum['payback_bidir_amb_deg'], 1),
                        round(resum['payback_difer_amb_deg'], 1),
                        round(resum['soc_min_year'] * 100, 1),
                        round(resum['total_co2'], 6)
                    ])
                    
                    tipus_hab = "PIS" if hab["tipus_consum"] == "PIS" else "CASA"
                    fv = "Si" if hab["FV_activada"] else "No"
                    mob_map = {"Esc1_Alt": "Alta", "Esc2_Mod": "Moderada", "Esc3_Bai": "Baixa"}
                    mobilitat = mob_map.get(mob["nom"], mob["nom"])
                    cext_map = {"Base": "Domestica", "CFein": "Feina", "CPubl": "Publica"}
                    carrega = cext_map.get(cext["nom"], cext["nom"])
                    
                    resultats_csv.append({
                        'habitatge': hab["nom"],
                        'tipus_hab': tipus_hab,
                        'fv': fv,
                        'mobilitat': mobilitat,
                        'carrega': carrega,
                        'd_lab': mob["d_lab"],
                        'd_fds': mob["d_fds"],
                        'estalvi_v2h': round(resum['total_estalvi'], 2),
                        'estalvi_fv': round(resum['total_estalvi_fv'], 2),
                        'estalvi_fv_ve': round(resum['total_estalvi_fv_ve'], 2),
                        'ingres_exc': round(resum['total_ingres_excedent'], 2),
                        'estalvi_total': round(resum['estalvi_total'], 2),
                        'cost_carg_ext': round(resum['total_cost_carg_ext'], 2),
                        'estalvi_net': round(resum['estalvi_net'], 2),
                        'factura_xarxa': round(resum['factura_anual'], 2),
                        'cost_total_energia': round(resum['cost_total_energia'], 2),
                        'estalvi_unidir': round(resum['estalvi_anual_unidir'], 2),
                        'v2h_kwh': round(resum['total_v2h'], 1),
                        'xarxa_kwh': round(resum['total_xarxa'], 1),
                        'carg_ext_kwh': round(resum['total_carg_ext'], 1),
                        'excedent_fv_kwh': round(resum['total_excedent'], 1),
                        'carg_fv_ve_kwh': round(resum['total_carg_fv_ve'], 1),
                        'demanda_kwh': round(resum['total_llar'], 1),
                        'autoconsum_pct': round(autoconsum * 100, 2),
                        'efc_v2h': round(resum['efc_v2h'], 1),
                        'cost_deg_anual': round(resum['cost_degradacio_anual'], 2),
                        'payback_bidir': round(resum['payback'], 1),
                        'payback_unidir': round(resum['payback_unidir'], 1),
                        'payback_difer': round(resum['payback_diferencial'], 1),
                        'pb_bidir_deg': round(resum['payback_bidir_amb_deg'], 1),
                        'pb_difer_deg': round(resum['payback_difer_amb_deg'], 1),
                        'soc_min_pct': round(resum['soc_min_year'] * 100, 1),
                        'tco2': round(resum['total_co2'], 6),
                    })
                    
                    comptador += 1

    print(f"\nGuardant TOT en un únic arxiu: {FITXER_SORTIDA} ...")
    wb.save(FITXER_SORTIDA)
    
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np
    
    print("\nGenerant gràfics de perfils de consum...")
    
    dt_cal = datetime.datetime(2025, 1, 1, 0, 0)
    cal_mes = []
    cal_hora = []
    cal_fds = []
    for i in range(8760):
        cal_mes.append(dt_cal.month)
        cal_hora.append(dt_cal.hour)
        cal_fds.append(1 if dt_cal.weekday() >= 5 else 0)
        dt_cal += datetime.timedelta(hours=1)
    
    cal_mes = np.array(cal_mes)
    cal_hora = np.array(cal_hora)
    cal_fds = np.array(cal_fds)
    arr_pis = np.array(llista_pis)
    arr_casa = np.array(llista_casa)
    arr_tarifes = np.array(llista_tarifes)
    arr_fv_total = np.array(llista_produccio_fv)
    arr_fv_auto = np.array(llista_autoconsum_fv)
    
    noms_mesos = ['Gen','Feb','Mar','Abr','Mai','Jun','Jul','Ago','Set','Oct','Nov','Des']
    colors_estacio = {'Hivern': '#3B8BD4', 'Primavera': '#5DCA5D', 'Estiu': '#EF9F27', 'Tardor': '#D85A30'}
    
    def estacio_mes(m):
        if m in [12, 1, 2]: return 'Hivern'
        if m in [3, 4, 5]: return 'Primavera'
        if m in [6, 7, 8]: return 'Estiu'
        return 'Tardor'
    
    fig, ax = plt.subplots(figsize=(10, 5))
    perfil_pis_h = [arr_pis[cal_hora == h].mean() for h in range(24)]
    perfil_casa_h = [arr_casa[cal_hora == h].mean() for h in range(24)]
    ax.plot(range(24), perfil_pis_h, 'o-', color='#3B8BD4', linewidth=2, label='Pis')
    ax.plot(range(24), perfil_casa_h, 's-', color='#D85A30', linewidth=2, label='Casa')
    ax.set_xlabel('Hora del dia')
    ax.set_ylabel('Consum mitjà (kWh/h)')
    ax.set_title('Perfil horari mitjà anual — Pis vs Casa')
    ax.set_xticks(range(24))
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig('Graf_Perfil_Horari_Anual.png', dpi=150)
    plt.close()
    print("  -> Graf_Perfil_Horari_Anual.png")
    
    fig, axes = plt.subplots(2, 2, figsize=(12, 8), sharey=True)
    fig.suptitle('Pis — Perfil horari per estació', fontsize=14)
    for idx, (est, col) in enumerate(colors_estacio.items()):
        ax = axes[idx // 2][idx % 2]
        mesos_est = [m for m in range(1, 13) if estacio_mes(m) == est]
        mask = np.isin(cal_mes, mesos_est)
        perfil = [arr_pis[mask & (cal_hora == h)].mean() for h in range(24)]
        ax.fill_between(range(24), perfil, alpha=0.3, color=col)
        ax.plot(range(24), perfil, '-', color=col, linewidth=2)
        ax.set_title(est)
        ax.set_xticks(range(0, 24, 3))
        ax.set_xlabel('Hora')
        ax.set_ylabel('kWh/h')
        ax.grid(True, alpha=0.3)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.savefig('Graf_Perfil_Estacional_PIS.png', dpi=150)
    plt.close()
    print("  -> Graf_Perfil_Estacional_PIS.png")

    fig, axes = plt.subplots(2, 2, figsize=(12, 8), sharey=True)
    fig.suptitle('Casa — Perfil horari per estació', fontsize=14)
    for idx, (est, col) in enumerate(colors_estacio.items()):
        ax = axes[idx // 2][idx % 2]
        mesos_est = [m for m in range(1, 13) if estacio_mes(m) == est]
        mask = np.isin(cal_mes, mesos_est)
        perfil = [arr_casa[mask & (cal_hora == h)].mean() for h in range(24)]
        ax.fill_between(range(24), perfil, alpha=0.3, color=col)
        ax.plot(range(24), perfil, '-', color=col, linewidth=2)
        ax.set_title(est)
        ax.set_xticks(range(0, 24, 3))
        ax.set_xlabel('Hora')
        ax.set_ylabel('kWh/h')
        ax.grid(True, alpha=0.3)
    plt.tight_layout(rect=[0, 0, 1, 0.95])
    plt.savefig('Graf_Perfil_Estacional_CASA.png', dpi=150)
    plt.close()
    print("  -> Graf_Perfil_Estacional_CASA.png")

    fig, ax = plt.subplots(figsize=(10, 5))
    consum_mensual_pis = [arr_pis[cal_mes == m].sum() for m in range(1, 13)]
    consum_mensual_casa = [arr_casa[cal_mes == m].sum() for m in range(1, 13)]
    x = np.arange(12)
    w = 0.35
    ax.bar(x - w/2, consum_mensual_pis, w, label='Pis', color='#3B8BD4', alpha=0.85)
    ax.bar(x + w/2, consum_mensual_casa, w, label='Casa', color='#D85A30', alpha=0.85)
    ax.set_xticks(x)
    ax.set_xticklabels(noms_mesos)
    ax.set_ylabel('Consum mensual (kWh)')
    ax.set_title('Consum mensual — Pis vs Casa')
    ax.legend()
    ax.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    plt.savefig('Graf_Consum_Mensual.png', dpi=150)
    plt.close()
    print("  -> Graf_Consum_Mensual.png")

    fig, axes = plt.subplots(1, 2, figsize=(14, 5), sharey=False)
    for idx, (nom, arr) in enumerate([('Pis', arr_pis), ('Casa', arr_casa)]):
        ax = axes[idx]
        perfil_lab = [arr[(cal_fds == 0) & (cal_hora == h)].mean() for h in range(24)]
        perfil_fds = [arr[(cal_fds == 1) & (cal_hora == h)].mean() for h in range(24)]
        ax.plot(range(24), perfil_lab, '-', color='#3B8BD4', linewidth=2, label='Laborable')
        ax.plot(range(24), perfil_fds, '--', color='#D85A30', linewidth=2, label='Cap de setmana')
        ax.set_title(f'{nom} — Laborable vs Cap de setmana')
        ax.set_xlabel('Hora')
        ax.set_ylabel('kWh/h')
        ax.set_xticks(range(0, 24, 3))
        ax.legend()
        ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig('Graf_Laborable_vs_FDS.png', dpi=150)
    plt.close()
    print("  -> Graf_Laborable_vs_FDS.png")

    fig, ax = plt.subplots(figsize=(10, 5))
    perfil_casa_h = [arr_casa[cal_hora == h].mean() for h in range(24)]
    perfil_fv_h = [arr_fv_total[cal_hora == h].mean() for h in range(24)]
    ax.fill_between(range(24), perfil_casa_h, alpha=0.3, color='#D85A30', label='Demanda Casa')
    ax.plot(range(24), perfil_casa_h, '-', color='#D85A30', linewidth=2)
    ax.fill_between(range(24), perfil_fv_h, alpha=0.3, color='#5DCA5D', label='Producció FV')
    ax.plot(range(24), perfil_fv_h, '-', color='#5DCA5D', linewidth=2)
    ax.set_xlabel('Hora del dia')
    ax.set_ylabel('kWh/h')
    ax.set_title('Casa — Demanda vs Producció FV (perfil horari mitjà anual)')
    ax.set_xticks(range(24))
    ax.legend()
    ax.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig('Graf_Demanda_vs_FV.png', dpi=150)
    plt.close()
    print("  -> Graf_Demanda_vs_FV.png")

    fig, ax = plt.subplots(figsize=(10, 4))
    perfil_tarifa = [arr_tarifes[cal_hora == h].mean() for h in range(24)]
    colors_tarifa = []
    for h in range(24):
        p = periode_tarifari(h, 0)  # Laborable
        if p == 'P1': colors_tarifa.append('#E24B4A')
        elif p == 'P2': colors_tarifa.append('#EF9F27')
        else: colors_tarifa.append('#5DCA5D')
    ax.bar(range(24), perfil_tarifa, color=colors_tarifa, alpha=0.85, edgecolor='white')
    ax.set_xlabel('Hora del dia')
    ax.set_ylabel('Tarifa mitjana (€/kWh)')
    ax.set_title('Perfil tarifari mitjà horari (laborable) — P1 Punta / P2 Pla / P3 Vall')
    ax.set_xticks(range(24))
    ax.grid(True, alpha=0.3, axis='y')
    from matplotlib.patches import Patch
    ax.legend(handles=[
        Patch(facecolor='#E24B4A', label='P1 Punta'),
        Patch(facecolor='#EF9F27', label='P2 Pla'),
        Patch(facecolor='#5DCA5D', label='P3 Vall'),
    ])
    plt.tight_layout()
    plt.savefig('Graf_Perfil_Tarifari.png', dpi=150)
    plt.close()
    print("  -> Graf_Perfil_Tarifari.png")

    fig, axes = plt.subplots(1, 2, figsize=(16, 6))
    for idx, (nom, arr) in enumerate([('Pis', arr_pis), ('Casa', arr_casa)]):
        ax = axes[idx]
        matriu = np.zeros((24, 12))
        for m in range(1, 13):
            for h in range(24):
                matriu[h, m-1] = arr[(cal_mes == m) & (cal_hora == h)].mean()
        im = ax.imshow(matriu, aspect='auto', cmap='YlOrRd', origin='lower')
        ax.set_xticks(range(12))
        ax.set_xticklabels(noms_mesos, fontsize=9)
        ax.set_yticks(range(0, 24, 3))
        ax.set_ylabel('Hora')
        ax.set_title(f'{nom} — Consum mitjà (kWh/h)')
        plt.colorbar(im, ax=ax, shrink=0.8)
    plt.tight_layout()
    plt.savefig('Graf_Heatmap_Consum.png', dpi=150)
    plt.close()
    print("  -> Graf_Heatmap_Consum.png")
    
    print("Gràfics de consum generats correctament.")
    
    print("\nGenerant fitxers CSV per anàlisi estadístic (Minitab)...")
    
    def fmt(val):
        """Formata un valor per CSV Minitab: números amb coma decimal, text sense canvis."""
        if isinstance(val, float):
            return str(val).replace('.', ',')
        elif isinstance(val, int):
            return str(val)
        return str(val)
    
    def escriure_csv(fitxer, capçalera, files):
        """Escriu CSV amb ; separador, coma decimal, UTF-8 amb BOM."""
        with open(fitxer, 'w', encoding='utf-8-sig') as f:
            f.write(';'.join(capçalera) + '\n')
            for fila in files:
                f.write(';'.join(fmt(v) for v in fila) + '\n')
    
    capçalera_factorial = [
        "Habitatge", "Tipus_Habitatge", "FV", "Mobilitat", "Carrega",
        "d_lab_km", "d_fds_km",
        "Estalvi_V2H_EUR", "Estalvi_FV_EUR", "Estalvi_FV_VE_EUR", "Ingres_Excedents_EUR",
        "Estalvi_Total_EUR", "Cost_Carrega_Ext_EUR", "Estalvi_Net_EUR",
        "Factura_Xarxa_EUR", "Cost_Total_Energia_EUR", "Estalvi_Unidir_EUR",
        "V2H_kWh", "Xarxa_Total_kWh", "Carrega_Ext_kWh", "Excedent_FV_kWh",
        "Carg_FV_VE_kWh",
        "Demanda_Llar_kWh", "Autoconsum_pct", "EFC_V2H",
        "Cost_Degradacio_EUR",
        "Payback_Bidir_anys", "Payback_Unidir_anys", "Payback_Diferencial_anys",
        "PB_Bidir_Deg_anys", "PB_Difer_Deg_anys",
        "SOC_min_pct", "tCO2_Anual",
    ]
    
    files_factorial = []
    for r in resultats_csv:
        files_factorial.append([
            r['habitatge'], r['tipus_hab'], r['fv'], r['mobilitat'], r['carrega'],
            r['d_lab'], r['d_fds'],
            r['estalvi_v2h'], r['estalvi_fv'], r['estalvi_fv_ve'], r['ingres_exc'],
            r['estalvi_total'], r['cost_carg_ext'], r['estalvi_net'],
            r['factura_xarxa'], r['cost_total_energia'], r['estalvi_unidir'],
            r['v2h_kwh'], r['xarxa_kwh'], r['carg_ext_kwh'], r['excedent_fv_kwh'],
            r['carg_fv_ve_kwh'],
            r['demanda_kwh'], r['autoconsum_pct'], r['efc_v2h'],
            r['cost_deg_anual'],
            r['payback_bidir'], r['payback_unidir'], r['payback_difer'],
            r['pb_bidir_deg'], r['pb_difer_deg'],
            r['soc_min_pct'], r['tco2'],
        ])
    
    fitxer_csv1 = "Minitab_Factorial.csv"
    escriure_csv(fitxer_csv1, capçalera_factorial, files_factorial)
    print(f"  -> {fitxer_csv1}: {len(files_factorial)} escenaris × {len(capçalera_factorial)} columnes")
    
    for cas, filtre in [("PIS", "PIS"), ("CASA", "CASA")]:
        files_cas = [f for f in files_factorial if f[1] == filtre]
        fitxer = f"Minitab_{cas}.csv"
        escriure_csv(fitxer, capçalera_factorial, files_cas)
        print(f"  -> {fitxer}: {len(files_cas)} escenaris")
    
    capçalera_payback = [
        "Habitatge", "FV", "Mobilitat", "Carrega",
        "Estalvi_Net_EUR", "Estalvi_FV_VE_EUR", "Estalvi_Unidir_EUR", "Cost_Degradacio_EUR",
        "Payback_Bidir_anys", "Payback_Unidir_anys", "Payback_Diferencial_anys",
        "PB_Bidir_Deg_anys", "PB_Difer_Deg_anys",
        "Bidir_Rendible_8a", "Unidir_Rendible_8a",
        "Bidir_Deg_Rendible_8a", "Recomanacio", "Reco_amb_Degradacio"
    ]

    files_payback = []
    for r in resultats_csv:
        pb_bidir = r['payback_bidir']
        pb_unidir = r['payback_unidir']
        pb_difer = r['payback_difer']
        pb_bidir_d = r['pb_bidir_deg']
        pb_difer_d = r['pb_difer_deg']

        bidir_ok = "Si" if pb_bidir < 8 else "No"
        unidir_ok = "Si" if pb_unidir < 8 else "No"
        bidir_deg_ok = "Si" if pb_bidir_d < 8 else "No"

        if pb_difer < 8:
            reco = "Bidireccional"
        elif pb_unidir < 8:
            reco = "Unidireccional"
        elif r['estalvi_net'] > 0:
            reco = "Unidir (bidir lent)"
        else:
            reco = "No rendible"

        if pb_difer_d < 8:
            reco_deg = "Bidireccional"
        elif pb_unidir < 8:
            reco_deg = "Unidireccional"
        elif r['estalvi_net'] > 0:
            reco_deg = "Unidir (bidir lent)"
        else:
            reco_deg = "No rendible"

        files_payback.append([
            r['tipus_hab'], r['fv'], r['mobilitat'], r['carrega'],
            r['estalvi_net'], r['estalvi_fv_ve'], r['estalvi_unidir'], r['cost_deg_anual'],
            pb_bidir, pb_unidir, pb_difer,
            pb_bidir_d, pb_difer_d,
            bidir_ok, unidir_ok, bidir_deg_ok,
            reco, reco_deg
        ])
    
    fitxer_csv4 = "Minitab_Decisio_Payback.csv"
    escriure_csv(fitxer_csv4, capçalera_payback, files_payback)
    print(f"  -> {fitxer_csv4}: {len(files_payback)} escenaris amb recomanació")
    
    print("\n" + "="*50)
    print("SIMULACIÓ FINALITZADA CORRECTAMENT")
    print("="*50)
    print(f"Resultats Excel: {FITXER_SORTIDA}")
    print(f"CSV Minitab:     {fitxer_csv1}, Minitab_PIS.csv, Minitab_CASA.csv, {fitxer_csv4}")